#!/usr/bin/env python3
"""Convert the approved bus sheets to privacy-minimized JSON for import review."""
from __future__ import annotations
import argparse, json, re
from datetime import time
from pathlib import Path
from openpyxl import load_workbook

BUS_SHEETS = ("1호차", "2호차", "3호차", "5호차", "6호차")
DAYS = {"월": "월", "화": "화", "수": "수", "목": "목", "금": "금"}

def clock(value):
    if isinstance(value, time): return value.strftime("%H:%M")
    if value is None: return ""
    match = re.search(r"(\d{1,2})\s*[:시]\s*(\d{1,2})", str(value))
    return f"{int(match.group(1)):02d}:{int(match.group(2)):02d}" if match else str(value).strip()

def slot_for(label):
    compact = str(label).replace(" ", "")
    if "9시" in compact and "등원" in compact: return "09-in"
    if "3시" in compact and "등원" in compact: return "15-in"
    if "3시" in compact and "하원" in compact: return "15-out"
    if "4시30분" in compact and "등원" in compact: return "1630-in"
    if "4시30분" in compact and "하원" in compact: return "1630-out"
    if "6시" in compact and "하원" in compact: return "18-out"

def split_name(value):
    raw = re.sub(r"\s+", " ", str(value or "")).strip()
    parts = raw.split(" ", 1)
    english = parts[1] if len(parts) == 2 and re.search(r"[A-Za-z]", parts[1]) else None
    return parts[0], english

def day_list(value):
    text = str(value or "").replace("매일", "월화수목금")
    return [day for key, day in DAYS.items() if key in text]

def merged_lookup(ws):
    """Resolve only layout cells such as vertically merged time/place values."""
    lookup={}
    for area in ws.merged_cells.ranges:
        anchor=ws.cell(area.min_row,area.min_col).value
        for row in range(area.min_row,area.max_row+1):
            for col in range(area.min_col,area.max_col+1):
                lookup[(row,col)]=anchor
    return lookup

def main():
    parser=argparse.ArgumentParser(); parser.add_argument("workbook"); parser.add_argument("output"); args=parser.parse_args()
    wb=load_workbook(args.workbook,data_only=True)
    rides=[]; review=[]
    for sheet in BUS_SHEETS:
        ws=wb[sheet]; bus=int(sheet[0]); current=None; split_layout=False
        merged=merged_lookup(ws)
        for row in range(1,ws.max_row+1):
            heading=ws.cell(row,2).value
            detected=slot_for(heading) if heading else None
            if detected:
                current=detected
                split_layout=str(ws.cell(row+1,8).value or "").strip()=="화목"
                continue
            if not current: continue
            place=str(merged.get((row,3),ws.cell(row,3).value) or "").strip()
            note=str(ws.cell(row,9).value or ws.cell(row,8).value or "").strip()
            weekday=day_list(ws.cell(row,6).value)
            left,right=ws.cell(row,5).value,ws.cell(row,8).value
            candidates=[]
            left_time=clock(merged.get((row,2),ws.cell(row,2).value))
            right_time=clock(merged.get((row,7),ws.cell(row,7).value))
            # Some split rows omit the second time when both sides share a stop/time.
            if not left_time: left_time=right_time
            if not right_time: right_time=left_time
            if not split_layout:
                # 9 a.m. and bus 6 are single-list sections; H contains notes.
                if left: candidates.append((left,left_time,weekday))
            else:
                if left and right:
                    left_days=[d for d in weekday if d in "월수금"]
                    right_days=[d for d in weekday if d in "화목"]
                    if left_days: candidates.append((left,left_time,left_days))
                    if right_days: candidates.append((right,right_time,right_days))
                elif left:
                    candidates.append((left,left_time,weekday))
                elif right:
                    candidates.append((right,right_time,weekday))
            for raw_name,ride_time,ride_days in candidates:
                student,english=split_name(raw_name)
                if not student or student in {"아동명","월수금"}: continue
                if not place or not ride_days:
                    review.append({"sheet":sheet,"row":row,"student":student,"reason":"시간·장소·요일 중 누락","note":note})
                    continue
                if not ride_time:
                    ride_time="미정"
                    review.append({"sheet":sheet,"row":row,"student":student,"reason":"시간 미입력","note":note})
                for day in ride_days:
                    rides.append({"id":f"{sheet}-{row}-{day}-{current}","bus":bus,"slot":current,"day":day,"time":ride_time,"stop":place,"student":student,"englishName":english if current.startswith("09") else None,"groupName":"오전유치부" if current.startswith("09") or current=="15-out" else "미분류","sourceSheet":sheet,"sourceRow":row})
            if note and any(token in note for token in ("부터","개별","요일","호차","변경")):
                review.append({"sheet":sheet,"row":row,"student":split_name(left or right)[0],"reason":"특이사항 수동 확인","note":note})
    Path(args.output).write_text(json.dumps({"rides":rides,"review":review,"excludedSheets":["학원","개별등하원 명단"]},ensure_ascii=False,indent=2),encoding="utf-8")
    print(json.dumps({"rides":len(rides),"review":len(review)},ensure_ascii=False))
if __name__ == "__main__": main()
